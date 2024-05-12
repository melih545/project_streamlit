import streamlit as st
import pandas as pd
from ydata_profiling import ProfileReport
from streamlit_ydata_profiling import st_profile_report

from prophet import Prophet
from prophet.plot import plot_plotly
import plotly.graph_objects as go
from openpyxl import load_workbook
from dateutil.relativedelta import relativedelta

from statsmodels.tsa.statespace.sarimax import SARIMAX

data_path = "project_copy.xlsm"

workbook = load_workbook(filename=data_path, read_only=True)
sheet = workbook['Düzenlenmiş Veri Tablosu']

lst  = []

for row in sheet.iter_rows(values_only=True):
    lst.append(row[:7])


df = pd.DataFrame(lst[1:], columns=lst[0])

season_dict = {
    'Ocak': 'Kış',
    'Şubat': 'Kış',
    'Mart': 'İlkbahar',
    'Nisan': 'İlkbahar',
    'Mayıs': 'İlkbahar',
    'Haziran': 'Yaz',
    'Temmuz': 'Yaz',
    'Ağustos': 'Yaz',
    'Eylül': 'Sonbahar',
    'Ekim': 'Sonbahar',
    'Kasım': 'Sonbahar',
    'Aralık': 'Kış'
}

month_dict = {
    'Ocak': 1,
    'Şubat': 2,
    'Mart': 3,
    'Nisan': 4,
    'Mayıs': 5,
    'Haziran': 6,
    'Temmuz': 7,
    'Ağustos': 8,
    'Eylül': 9,
    'Ekim': 10,
    'Kasım': 11,
    'Aralık': 12
}

# sidebar
option = st.sidebar.radio('Select...',
                          ['pandas_profiling', 'forecast'])

if option == 'pandas_profiling':
    # description
    st.title('Altınay Defense Technolojileri Gelir Gider Analizi')

    st.subheader('First look into the data with pandas_profiling')
    selected_table = st.selectbox("What file do you want to see first?",
                                  [data_path])

    st.write(df.head())
    st.write(df.dtypes)

    olist_profile = ProfileReport(df, explorative=True)
    st_profile_report(olist_profile)

      
if option == 'forecast':
    model = Prophet(yearly_seasonality = True, seasonality_prior_scale=0.1)
    df["Tarih"] = pd.to_datetime(df['Yıl'].astype(str) + '-' + df['Ay'].map(month_dict).astype(str), infer_datetime_format=True)
    
    starting_date = df['Tarih'].max()
    last_date = st.date_input("Hangi tarihe kadar forecast etmek istiyorsunuz?", value=starting_date)
    years = relativedelta(last_date, starting_date).years
    month = relativedelta(last_date, starting_date).months
    months = (12*years) + month

    gelir_rows = df[df["Tür"] == "Gelir"]
    gelir_rows = gelir_rows.groupby(['Tarih', 'Tür'], as_index=False)['Miktar'].sum()
    gelir_rows['ds'] = gelir_rows['Tarih']
    gelir_rows['y'] = gelir_rows['Miktar']

    model.fit(gelir_rows[['ds', 'y']])

    future = model.make_future_dataframe(periods=months, freq='MS')

    forecast = model.predict(future)

    fig3 = plot_plotly(model, forecast)
    fig3.update_traces(marker=dict(color="white"))

    fig3.update_layout(
        title="Gelir Tahmini",
        xaxis_title="Tarih",
        yaxis_title="Gelir",
        template="plotly_dark", 
        showlegend=True,  
        hovermode="x")

    st.write(fig3)

    fig2 = model.plot_components(forecast)
    st.write(fig2)

