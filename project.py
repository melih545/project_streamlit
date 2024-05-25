import pandas as pd
import streamlit as st
import plotly.graph_objs as go
from ydata_profiling import ProfileReport
from streamlit_ydata_profiling import st_profile_report

from prophet import Prophet
from openpyxl import load_workbook
from prophet.plot import plot_plotly
from dateutil.relativedelta import relativedelta


try:
    data_path = "project.xlsm"

except:
    data_path = "/mnt/c/Users/edipm/Desktop/Ders/3. Sınıf/Advanced Spreadsheet/Project/project.xlsm"

workbook = load_workbook(filename=data_path, read_only=True)
sheet = workbook["Düzenlenmiş Veri Tablosu"]

lst = []

for row in sheet.iter_rows(values_only=True):
    lst.append(row[:7])


df = pd.DataFrame(lst[1:], columns=lst[0])

season_dict = {
    "Ocak": "Kış",
    "Şubat": "Kış",
    "Mart": "İlkbahar",
    "Nisan": "İlkbahar",
    "Mayıs": "İlkbahar",
    "Haziran": "Yaz",
    "Temmuz": "Yaz",
    "Ağustos": "Yaz",
    "Eylül": "Sonbahar",
    "Ekim": "Sonbahar",
    "Kasım": "Sonbahar",
    "Aralık": "Kış",
}

month_dict = {
    "Ocak": 1,
    "Şubat": 2,
    "Mart": 3,
    "Nisan": 4,
    "Mayıs": 5,
    "Haziran": 6,
    "Temmuz": 7,
    "Ağustos": 8,
    "Eylül": 9,
    "Ekim": 10,
    "Kasım": 11,
    "Aralık": 12,
}
month_num_dict = {
    1: "Ocak",
    2: "Şubat",
    3: "Mart",
    4: "Nisan",
    5: "Mayıs",
    6: "Haziran",
    7: "Temmuz",
    8: "Ağustos",
    9: "Eylül",
    10: "Ekim",
    11: "Kasım",
    12: "Aralık"
}


option = st.sidebar.radio("Seçenekler", ["Analiz", "Filtreleme", "Veri Tahmini"])

if option == "Analiz":
    st.title("Altınay Savunma Teknolojileri Gelir Gider Analizi")

    st.subheader("\nHangi veriyi görmek istiyorsunuz?")
    selected_table = st.selectbox("Hangi veriyi görmek istiyorsunuz?", ["Gelir Gider Verileri"], label_visibility="hidden")

    st.write(df)
    st.write(df.dtypes)

    olist_profile = ProfileReport(df, explorative=True)
    st_profile_report(olist_profile)


if option == "Veri Tahmini":
    df["Tarih"] = pd.to_datetime(
        df["Yıl"].astype(str) + "-" + df["Ay"].map(month_dict).astype(str),
        infer_datetime_format=True
    )
    forecast_type = st.selectbox(
        "Ne tahmini yapmak istiyorsunuz", options=["Gelir", "Gider", "Kâr", "Gelir_Gider"]
    )
    starting_date = df["Tarih"].max()
    last_date = st.date_input(
        "Hangi tarihe kadar tahmin etmek istiyorsunuz?", value=starting_date
    )

    years = relativedelta(last_date, starting_date).years
    month = relativedelta(last_date, starting_date).months
    months = (12 * years) + month

    if forecast_type == "Gelir":
        table_name = "Gelir Tahmini"
        y_name = "Gelir"
        model = Prophet(yearly_seasonality=True, seasonality_prior_scale=0.1, seasonality_mode='multiplicative')
        forecast_df = df[df["Tür"] == "Gelir"]
        forecast_df = forecast_df.groupby(["Tarih", "Tür"], as_index=False)["Miktar"].sum()

    elif forecast_type == "Gider":
        table_name = "Gider Tahmini"
        y_name = "Gider"
        model = Prophet(yearly_seasonality=True, seasonality_prior_scale=0.1)
        forecast_df = df[df["Tür"] == "Gider"]
        forecast_df = forecast_df.groupby(["Tarih", "Tür"], as_index=False)["Miktar"].sum()

    elif forecast_type == "Kâr":
        model = Prophet(yearly_seasonality=True, seasonality_prior_scale=0.1)
        forecast_df = df
        forecast_df.loc[forecast_df["Tür"] == "Gider", "Miktar"] *= -1
        forecast_df = forecast_df.groupby(["Tarih"], as_index=False)["Miktar"].sum()

    if forecast_type == "Gelir_Gider":
        gelir_model = Prophet(yearly_seasonality=True, seasonality_prior_scale=0.1)
        gider_model = Prophet(yearly_seasonality=True, seasonality_prior_scale=0.1)

        gelir_df = df[df["Tür"] == "Gelir"]
        gelir_df = gelir_df.groupby(["Tarih", "Tür"], as_index=False)["Miktar"].sum()

        gider_df = df[df["Tür"] == "Gider"]
        gider_df = gider_df.groupby(["Tarih", "Tür"], as_index=False)["Miktar"].sum()

        gelir_df["ds"] = gelir_df["Tarih"]
        gelir_df["y"] = gelir_df["Miktar"]

        gider_df["ds"] = gider_df["Tarih"]
        gider_df["y"] = gider_df["Miktar"]

        gelir_model.fit(gelir_df[["ds", "y"]])
        gider_model.fit(gider_df[["ds", "y"]])
        
        gelir_future = gelir_model.make_future_dataframe(periods=months, freq="MS")
        gider_future = gider_model.make_future_dataframe(periods=months, freq="MS")

        gelir_forecast = gelir_model.predict(gelir_future)
        gider_forecast = gider_model.predict(gider_future)

        gelir_forecast = gelir_forecast[gelir_forecast['ds'] >= starting_date]
        gider_forecast = gider_forecast[gider_forecast['ds'] >= starting_date]
        
        trace1 = go.Scatter(x=gelir_forecast['ds'], y=gelir_forecast['yhat'], mode='lines', name='Gelir Tahmini')
        trace2 = go.Scatter(x=gider_forecast['ds'], y=gider_forecast['yhat'], mode='lines', name='Gider Tahmini')

        trace3 = go.Scatter(x=gelir_df['Tarih'], y=gelir_df['Miktar'], mode='lines', name='Gelir')
        trace4 = go.Scatter(x=gider_df['Tarih'], y=gider_df['Miktar'], mode='lines', name='Gider')

        data = [trace1, trace3, trace2, trace4]
        layout = go.Layout(title='Gelir-Gider Tahmin Tablosu', xaxis=dict(title='Tarih'), yaxis=dict(title='Miktar'))

        fig = go.Figure(data=data, layout=layout)

        st.plotly_chart(fig)

    else:
        forecast_df["ds"] = forecast_df["Tarih"]
        forecast_df["y"] = forecast_df["Miktar"]

        model.fit(forecast_df[["ds", "y"]])

        future = model.make_future_dataframe(periods=months, freq="MS")
        forecast = model.predict(future)
        
        if forecast_type == "Kâr":
            table_name = "Kâr Tahmini"
            y_name = "Kâr"
            below_zero = forecast.loc[forecast["yhat"] <= 0]

            st.header("Zararda olduğumuz aylar ve zarar miktarları")
            if below_zero.empty:
                st.write("Tahminlere göre zarara geçtiğimiz bir ay yok!")

            else:
                below_zero = below_zero.rename(columns={"ds": "Tarih", "yhat": "Zarar Miktarı"})
                below_zero["Zarar Miktarı"] = below_zero["Zarar Miktarı"].astype(int)
                below_zero['Yıl'] = below_zero['Tarih'].dt.year.astype(str)
                below_zero['Ay'] = below_zero['Tarih'].dt.month.map(month_num_dict).astype(str)
                st.write(below_zero[["Yıl", "Ay", "Zarar Miktarı"]])    


        fig3 = plot_plotly(model, forecast)
        fig3.update_traces(marker=dict(color="white"))

        fig3.update_layout(
            title=table_name,
            xaxis_title="Tarih",
            yaxis_title=y_name,
            template="plotly_dark",
            showlegend=True,
            hovermode="x",
        )

        st.write(fig3)

        fig2 = model.plot_components(forecast)
        st.write(fig2)

if option == "Filtreleme":
    kalem_col, altkalem_col, açıklama_col = st.columns([5, 5, 5])

    with kalem_col:
        kalem = st.selectbox("Kalem",  [" "] + df["Kalem"].unique().tolist())
        if kalem != " ":
            df = df[df["Kalem"] == kalem]

    with altkalem_col:
        alt_kalem = st.selectbox("Alt Kalem", [" "] + df["Alt Kalem"].unique().tolist())
        if alt_kalem != " ":
            df = df[df["Alt Kalem"] == alt_kalem]
        
    with açıklama_col:
        açıklama = st.selectbox("İlgili Açıklama", [" "] + df["İlgili / Açıklama"].unique().tolist())
        if açıklama != " " and açıklama is not None :
            df = df[df["İlgili / Açıklama"] == açıklama]

    miktar_col, ay_col, yıl_col, tür_col = st.columns([5, 5, 5, 5])

    with miktar_col:
        miktar = st.selectbox("Miktar", [" "] + df["Miktar"].unique().tolist())
        if miktar != " ":
            df = df[df["Miktar"] == miktar]
            
    with ay_col:
        ay = st.selectbox("Ay", [" "] + df["Ay"].unique().tolist())
        if ay != " ":
            df = df[df["Ay"] == ay]

    with yıl_col:
        yıl = st.selectbox("Yıl", [" "] + df["Yıl"].unique().tolist())
        if yıl != " ":
            df = df[df["Yıl"] == yıl]
        
    with tür_col:
        tür = st.selectbox("Tür", [" "] + df["Tür"].unique().tolist())
        if tür != " ":
            df = df[df["Tür"] == tür]

    row_list = [["Kalem", kalem], ["Alt Kalem", alt_kalem], ["İlgili / Açıklama", açıklama], ["Miktar", miktar], 
                ["Ay", ay], ["Yıl", yıl], ["Tür", tür]]

    if st.button("Filtrele"):
        st.write(df, width=1000)
        
