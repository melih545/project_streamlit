import streamlit as st
import pandas as pd
from ydata_profiling import ProfileReport
from streamlit_ydata_profiling import st_profile_report

import numpy as np

import pmdarima as pm
from prophet import Prophet
from prophet.plot import plot_plotly
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import plotly.graph_objects as go
import datetime as dt
from openpyxl import load_workbook

from statsmodels.tsa.statespace.sarimax import SARIMAX

data_path = "/mnt/c/Users/edipm/Desktop/Ders/3. Sınıf/Advanced Spreadsheet/Project/project_copy.xlsm"

workbook = load_workbook(filename=data_path, read_only=True)
sheet = workbook['Düzenlenmiş Veri Tablosu']

lst  = []

for row in sheet.iter_rows(values_only=True):
    lst.append(row[:7])


df = pd.DataFrame(lst[1:], columns=lst[0])

season_dict = {
    'Ocak': 'Kış',
    'Şubat': 'Kış',
    'Mart': 'İlkbahar',
    'Nisan': 'İlkbahar',
    'Mayıs': 'İlkbahar',
    'Haziran': 'Yaz',
    'Temmuz': 'Yaz',
    'Ağustos': 'Yaz',
    'Eylül': 'Sonbahar',
    'Ekim': 'Sonbahar',
    'Kasım': 'Sonbahar',
    'Aralık': 'Kış'
}

month_dict = {
    'Ocak': 1,
    'Şubat': 2,
    'Mart': 3,
    'Nisan': 4,
    'Mayıs': 5,
    'Haziran': 6,
    'Temmuz': 7,
    'Ağustos': 8,
    'Eylül': 9,
    'Ekim': 10,
    'Kasım': 11,
    'Aralık': 12
}

def sarimax_forecast(SARIMAX_model, periods=12):
    # Forecast
    n_periods = periods
    print("predict'e başlıyoz")
 
    forecast_df = pd.DataFrame({"month_index": pd.date_range(df.index[-1], periods=n_periods, freq='MS').month},
                               index=pd.date_range(df.index[-1] + pd.DateOffset(months=1), periods=n_periods, freq='MS'))
 
    fitted, confint = SARIMAX_model.predict(n_periods=n_periods,
                                            return_conf_int=True,
                                            exogenous=forecast_df[['month_index']])
    index_of_fc = pd.date_range(df.index[-1] + pd.DateOffset(months=1), periods=n_periods, freq='MS')
    print("predict bitti")
    # make series for plotting purpose
    fitted_series = pd.Series(fitted, index=index_of_fc)
    lower_series = pd.Series(confint[:, 0], index=index_of_fc)
    upper_series = pd.Series(confint[:, 1], index=index_of_fc)
 
    fig = go.Figure()

    # Gerçek miktarın çizimi
    fig.add_trace(go.Scatter(x=df.index, y=df["Miktar"], mode='lines', name='Gerçek Miktar', line=dict(color='#1f76b4')))

    # Uygunlaştırılmış serinin çizimi
    fig.add_trace(go.Scatter(x=fitted_series.index, y=fitted_series, mode='lines', name='Uygunlaştırılmış Seri', line=dict(color='darkgreen')))

    # Güven aralığının çizimi
    fig.add_trace(go.Scatter(x=lower_series.index, y=lower_series, mode='lines', name='Alt Sınır', line=dict(color='black'), fill=None))
    fig.add_trace(go.Scatter(x=upper_series.index, y=upper_series, mode='lines', name='Üst Sınır', line=dict(color='black'), fill='tonexty'))

    # Grafik düzenlemeleri
    fig.update_layout(
        xaxis_title='Tarih',
        yaxis_title='Miktar',
        title='SARIMAX - Havayolu Yolcu Tahmini'
    )

    # Grafiği Streamlit'e gömme
    st.plotly_chart(fig)

# sidebar
option = st.sidebar.radio('Select...',
                          ['pandas_profiling', 'forecast'])

if option == 'pandas_profiling':
    # description
    st.title('Altınay Defense Technolojileri Gelir Gider Analizi')

    st.subheader('First look into the data with pandas_profiling')
    selected_table = st.selectbox("What file do you want to see first?",
                                  [data_path])

    st.write(df.head())
    st.write(df.dtypes)

    olist_profile = ProfileReport(df, explorative=True)
    st_profile_report(olist_profile)

      
if option == 'forecast':
    # Prophet modelini oluşturalım ve eğitelim
    starting_date = st.date_input("Forecast başlangıç tarihi")
    last_date = st.date_input("Hangi tarihe kadar forecast etmek istiyorsunuz?")
    
    model = Prophet(yearly_seasonality = True, seasonality_prior_scale=0.1)
    df["Tarih"] = pd.to_datetime(df['Yıl'].astype(str) + '-' + df['Ay'].map(month_dict).astype(str), infer_datetime_format=True)
    gelir_rows = df[df["Tür"] == "Gelir"]
    gelir_rows['ds'] = gelir_rows['Tarih']
    gelir_rows['y'] = gelir_rows['Miktar']
    print(gelir_rows[['Tarih', "Miktar"]])
    model.fit(gelir_rows[['ds', 'y']])

    # 12 aylık bir tahmin oluşturalım
    future = model.make_future_dataframe(periods=12, freq='MS')

    # Tahmin yapalım
    forecast = model.predict(future)

    # Tahmin sonuçlarını gösterelim
    print(forecast[['ds', 'yhat', 'yhat_lower', 'yhat_upper']].tail(12))

    # Tahminleri Plotly ile görselleştirin
    fig3 = plot_plotly(model, forecast)
    fig3.update_traces(marker=dict(color="white"))
    # Grafik özelliklerini ayarlayın
    fig3.update_layout(
        title="Gelir Tahmini",
        xaxis_title="Tarih",
        yaxis_title="Gelir",
        template="plotly_dark",  # Tema
        showlegend=True,  # Açıklamayı göster
        hovermode="x",  # Fareyle üstüne gelindiğinde tüm verileri göster
    )

    st.write(fig3)

    # m = Prophet(yearly_seasonality = True, seasonality_prior_scale=0.1)
    # m.fit(gelir_rows[['ds', 'y']])
    # future = m.make_future_dataframe(periods=12)
    # forecast = m.predict(future)
    fig2 = model.plot_components(forecast)
    st.write(fig2)

